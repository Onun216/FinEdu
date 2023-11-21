import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from finEdu.files_portfolio.excel_info import dividend_color, enter_position_color, close_position_color

file = load_workbook('/Users/nuno/FinEdu/finEdu/files_portfolio/Investimentos.xlsx', data_only=True)
file_sheets = file.sheetnames

# sector_positions = {
# 'ETF': [],
# 'Semiconductors': [],
# 'Basic Materials': [],
# 'Financial': [],
# 'Services': [],
# 'Consumer Cyclical': [],
# 'Goods': [],
# 'Transportation': [],
# 'Technology': [],
# 'REIT': [],
# 'Utilities': [],
# 'Energy': []
# }


def get_dividends(): 
    year = 2022
    total_dividends = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        dividends = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if bgColor == dividend_color:
                    dividends.append(cell_value)
                total = sum(dividends)
        total_dividends.append(total)
        #print(f'Dividendos de %s: €{total:.2f}' % ano)
        year += 1
    #print(f'Total: €{round(sum(total_dividends), 2)}')
    #print('------------------------------------')
    return total_dividends

def dividends_per_year(func, year):
    dividends_dict = {}
    for i, dividend in enumerate(func):
        dividends_dict[i + year] = round(dividend, 2)
    return dividends_dict

def dividends_growth(dividends):
    dividends = dividends_per_year(get_dividends(), 2022)
    previous_year_amount = None
    for year, amount in dividends.items():
        yoy = None
        if previous_year_amount is not None:
            yoy = (amount - previous_year_amount) / previous_year_amount * 100
        dividends[year] = {'amount': amount, 'YoY': yoy}
        previous_year_amount = amount
    return dividends





def get_positions():
    year = 2022
    total_positions = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        position = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if fs.cell(column, 4).fill.bgColor.index == dividend_color:
                    continue
                elif fs.cell(column, 4).fill.bgColor.index == close_position_color:
                    continue
                elif isinstance(cell_value, float):
                    position.append(cell_value)
                total = round(sum(position), 2)
        total_positions.append(total)
        #print(f'Posição em %s: €{total_position}' % ano)
        year += 1

    #print(f'Total: €{round(sum(total_position), 2)}')
    return total_positions


#cash_position = input("Insira o valor actual disponível para investir: ")


def total_positions_and_dividends():
    year = 2022
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        dividends = []
        position = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if not isinstance(cell_value, float):
                    continue
                # if bgColor == close_position_color:
                # continue
                if bgColor == dividend_color:
                    dividends.append(cell_value)
                total_div = sum(dividends)
                if isinstance(cell_value, float) and not (bgColor == dividend_color):
                    position.append(cell_value)
                total_pos = round(sum(position), 2)

        total_dividends.append(total_div)
        #print(f'Dividendos de %s: €{total_div:.2f}' % ano)

        # total_position.append(total_pos)
        # print(f'Posição em %s: €{total_position}' % ano)
        # print(position)

        year += 1
    # print(f'Custo actual das posições: €{round(sum(total_position), 2)}')
    #print(f'Total (dividendos): €{round(sum(total_dividends), 2)}')
    #print(f'Total (CASH): €{cash_position}')
    #print('------------------------------------')



