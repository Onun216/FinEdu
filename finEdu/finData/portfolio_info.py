import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import yfinance as yf
import requests

file = load_workbook('/Users/nuno/FinEdu/finEdu/files/Empresas_GPS.xlsx', data_only=True)
file_sheets = file.sheetnames

# lons, lats = [], []

def get_portfolio_constituents():
    names = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column

        for row in range(2, fs_count_row + 1):
            for column in range(1, 2):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                names.append(cell_value)

        return names

def get_hq_address():
    hq_address = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column

        for row in range(2, fs_count_row + 1):
            for column in range(2, 3):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                details.append(cell_value)

        return hq_address
    

constituents = get_portfolio_constituents()
# addresses = get_hq_address()





# tickers = get_ticker_symbols(constituents)

def verify_ticker_symbols(list):
    ticker_symbols = []
    for ticker in list:
        try:
            ticker_symbol = yf.Ticker(ticker).info
            ticker_symbols.append(ticker_symbol)
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                print(f"Ticker symbol '{ticker}' not found.")
            else:
                raise e
    
    return ticker_symbols


